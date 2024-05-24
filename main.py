from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
from tkinter import filedialog
import customtkinter
import tkinter as tk
from PIL import Image
from tkcalendar import DateEntry
from scrollframe import ScrollableCheckBoxFrame
from scrollframe import ScrollableCheckBoxFrame1
from read import readFile
from read import createFile
from openpyxl import Workbook
from openpyxl import load_workbook
from tkinter.messagebox import showerror, showwarning, showinfo

customtkinter.set_appearance_mode("light")


class ToplevelWindow(customtkinter.CTkToplevel):
    
    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        self.title('Amrita OD Managment System')
        self.state('zoomed')
        self.minsize(1200, 600)
        self.wm_iconbitmap(bitmap='icon.ico')
        self.customfont = customtkinter.CTkFont(
            family='Corbal', size=15, weight='normal')
        self.file_path = readFile('path.txt')
        self.wb = load_workbook(self.file_path)
        self.ws = self.wb.active
        self.excel_data = [(item[1].value, item[2].value)
                            for item in self.ws.iter_rows(min_row=2)]

        self.form_frame = FormFrame(self)
        self.form_frame.pack(fill='y', side='left', ipadx=30, ipady=20)
        self.list_frame = ListFrame(self, fg_color='gray92')
        self.list_frame.pack(fill='y', side='right')

        self.amrita = customtkinter.CTkImage(
            light_image=Image.open("logo.png"), size=(300, 74))
        self.amrita_logo = customtkinter.CTkLabel(
            master=self.form_frame, image=self.amrita, text='')
        self.amrita_logo.grid(row=0, column=0, columnspan=2, pady=20, padx=20)

        self.event_label = customtkinter.CTkLabel(
            master=self.form_frame, text='Event: ', compound='right', anchor='w', font=self.customfont)
        self.event_label.grid(row=1, column=0, padx=20, sticky='e')

        self.event_entry = customtkinter.CTkEntry(
            master=self.form_frame, placeholder_text='eg: Kalanjali', width=200)
        self.event_entry.grid(row=1, column=1, padx=20)

        self.from_label = customtkinter.CTkLabel(
            master=self.form_frame, text='From Date: ', compound='right', anchor='w', font=self.customfont)
        self.from_label.grid(row=2, column=0, padx=20, pady=10, sticky='e')

        self.from_entry = DateEntry(
            master=self.form_frame, date_format='d-m-yy', width=30)
        self.from_entry.grid(row=2, column=1, padx=20, pady=10)

        self.to_label = customtkinter.CTkLabel(
            master=self.form_frame, text='To Date: ', compound='right', anchor='w', font=self.customfont)
        self.to_label.grid(row=3, column=0, padx=20, pady=10, sticky='e')

        self.to_entry = DateEntry(
            master=self.form_frame, date_format='d-m-yy', width=30)
        self.to_entry.grid(row=3, column=1, padx=20, pady=10)

        self.session_label = customtkinter.CTkLabel(
            master=self.form_frame, text='Session: ', compound='right', anchor='w', font=self.customfont)
        self.session_label.grid(row=4, column=0, padx=20, pady=10, sticky='e')

        optionmenu_var = customtkinter.StringVar(value="Forenoon")
        self.session_entry = customtkinter.CTkOptionMenu(master=self.form_frame, values=[
                                                            "Forenoon", "Afternoon", "AllDay", "Other"], variable=optionmenu_var, width=200)
        self.session_entry.grid(row=4, column=1, padx=20, pady=10)

        self.hours_label = customtkinter.CTkLabel(
            master=self.form_frame, text='Hours: ', compound='right', anchor='w', font=self.customfont)
        self.hours_label.grid(row=5, column=0, padx=20, pady=10, sticky='e')

        self.hours_entry = customtkinter.CTkEntry(
            master=self.form_frame, placeholder_text='2-5', width=200)
        self.hours_entry.grid(row=5, column=1, padx=20, pady=10)

        self.remarks_label = customtkinter.CTkLabel(
            master=self.form_frame, text='Remarks: ', compound='right', anchor='w', font=self.customfont)
        self.remarks_label.grid(row=6, column=0, padx=20, pady=10, sticky='e')

        self.remarks_entry = customtkinter.CTkEntry(
            master=self.form_frame, placeholder_text='remarks', width=200)
        self.remarks_entry.grid(row=6, column=1, padx=20, pady=10)

        self.list_items = [f"{item[0]}   {item[1]}" for item in self.excel_data]
        self.listbox = ScrollableCheckBoxFrame(
            self.list_frame, label_text="List of students", width=350, height=500, item_list=self.list_items, label_fg_color='#00ADFE')
        self.listbox.grid(row=0, column=0, padx=10, pady=30, columnspan=2)

        self.sel_listbox = ScrollableCheckBoxFrame1(
            self.list_frame, width=350, height=500, label_text="Selected students", label_fg_color='#00ADFE')
        self.sel_listbox.grid(row=0, column=2, padx=10, pady=10, columnspan=2)

        self.students_label = customtkinter.CTkLabel(
            self.list_frame, text=f'Total count:   {len(self.list_items)}')
        self.students_label.grid(row=1, column=0, columnspan=2)

        self.selected_label = customtkinter.CTkLabel(
            self.list_frame, text=f'Total count:   {len(self.sel_listbox.get_checked_items())}')
        self.selected_label.grid(row=1, column=2, columnspan=2)

        # self.select_all = customtkinter.CTkButton(
        #     self.list_frame, text='Select all', height=32, command=self.test)
        # self.select_all.grid(row=2, column=0, padx=10, pady=10)

        self.push_selected = customtkinter.CTkButton(
            self.list_frame, text='Confirm', height=32, command=self.push_selected)
        self.push_selected.grid(row=2, column=0, padx=10, pady=10,columnspan=2)

        self.clear_all = customtkinter.CTkButton(
            self.list_frame, text='Clear all', height=32, command=self.reset_all)
        self.clear_all.grid(row=2, column=2, padx=10, pady=10)

        self.save = customtkinter.CTkButton(
            self.list_frame, text='Save', fg_color='green', height=32, command=self.save_all)
        self.save.grid(row=2, column=3, padx=10, pady=10)

    def reset_all(self):
        students = self.sel_listbox.get_checked_items()
        if students:
            for student in students:
                self.sel_listbox.remove_item(f'{student}')
                self.selected_label.configure(
                    text=f'Total count:   {len(self.sel_listbox.get_checked_items())}')

    # def test(self):
    #     self.listbox.select_allc = True
    #     print(self.listbox.select_allc)

    def push_selected(self):
        print(self.listbox.get_unchecked_items)
        students = self.listbox.get_checked_items()
        if students:
            for student in students:
                self.sel_listbox.add_item(f'{student}')
                self.selected_label.configure(
                    text=f'Total count:   {len(self.sel_listbox.get_checked_items())}')

    def save_all(self):
        students = self.sel_listbox.get_checked_items()
        reg = []
        stu = []
        for i in range(len(students)):
            splice = students[i].split('   ')[0]
            reg.append(splice)
            splice2 = students[i].split('   ')[1]
            stu.append(splice2)

        result = list(zip(reg, stu))

        self.event_name = self.event_entry.get()
        self.from_date = self.from_entry.get_date()
        self.to_date = self.to_entry.get_date()
        self.session = self.session_entry.get()
        self.hours = self.hours_entry.get()
        self.remarks = self.remarks_entry.get()
        if self.session == 'Allday':
            self.hours = 'NA'

        if(self.event_name and self.session and self.hours and result):
            # Create a new workbook if it doesn't exist
            user_filename = filedialog.asksaveasfilename(
                defaultextension='.xlsx')
            if user_filename:
                try:
                    wbf = load_workbook(user_filename)
                except:
                    wbf = Workbook()

                wsf = wbf.active
                wsf['A1'] = 'Event Name'
                wsf['B1'] = 'From Date'
                wsf['C1'] = 'To Date'
                wsf['D1'] = 'Reg.No'
                wsf['E1'] = 'Student Name'
                wsf['F1'] = 'Session'
                wsf['G1'] = 'No. of Hours'
                wsf['H1'] = 'Remarks'
                alignment = Alignment(
                    wrap_text=True, horizontal='center', vertical='center')

                # Append data to the next empty row
                for i, (key, value) in enumerate(result):
                    wsf.append([self.event_name, self.from_date, self.to_date,
                                key, value, self.session, self.hours, self.remarks])
                    wbf.save(user_filename)
                showinfo(
                    title='Succuss',
                    message='Data has been saved successfully!')
            else:
                showerror(title='Error', message="File couldn't be saved")
        else:
            showerror(title='Error', message='Please fill all the Entries')

class FormFrame(customtkinter.CTkFrame):
    def __init__(self, master, **kwargs):
        super().__init__(master, **kwargs)


class ListFrame(customtkinter.CTkFrame):
    def __init__(self, master, **kwargs):
        super().__init__(master, **kwargs)


class HomePage(customtkinter.CTk):
    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        self.title('Amrita OD Managment System')
        self.minsize(1200,600)
        self.iconbitmap(bitmap='icon.ico')
        self.customfont = customtkinter.CTkFont(
            family='Corbal', size=15, weight='normal')
        self.file_path = readFile('path.txt')
        self.wb = load_workbook(self.file_path)
        self.ws = self.wb.active
        self.excel_data = [(item[1].value, item[2].value)
                           for item in self.ws.iter_rows(min_row=2)]

        self.form_frame = FormFrame(self)
        self.form_frame.pack(fill='y', side='left', ipadx=30, ipady=20)
        self.list_frame = ListFrame(self, fg_color='gray92')
        self.list_frame.pack(fill='y', side='right')

        self.amrita = customtkinter.CTkImage(
            light_image=Image.open("logo.png"), size=(300, 74))
        self.amrita_logo = customtkinter.CTkLabel(
            master=self.form_frame, image=self.amrita, text='')
        self.amrita_logo.grid(row=0, column=0, columnspan=2, pady=20, padx=20)

        self.event_label = customtkinter.CTkLabel(
            master=self.form_frame, text='Event: ', compound='right', anchor='w', font=self.customfont)
        self.event_label.grid(row=1, column=0, padx=20, sticky='e')

        self.event_entry = customtkinter.CTkEntry(
            master=self.form_frame, placeholder_text='eg: Kalanjali', width=200)
        self.event_entry.grid(row=1, column=1, padx=20)

        self.from_label = customtkinter.CTkLabel(
            master=self.form_frame, text='From Date: ', compound='right', anchor='w', font=self.customfont)
        self.from_label.grid(row=2, column=0, padx=20, pady=10, sticky='e')

        self.from_entry = DateEntry(
            master=self.form_frame, date_format='d-m-yy', width=30)
        self.from_entry.grid(row=2, column=1, padx=20, pady=10)

        self.to_label = customtkinter.CTkLabel(
            master=self.form_frame, text='To Date: ', compound='right', anchor='w', font=self.customfont)
        self.to_label.grid(row=3, column=0, padx=20, pady=10, sticky='e')

        self.to_entry = DateEntry(
            master=self.form_frame, date_format='d-m-yy', width=30)
        self.to_entry.grid(row=3, column=1, padx=20, pady=10)

        self.session_label = customtkinter.CTkLabel(
            master=self.form_frame, text='Session: ', compound='right', anchor='w', font=self.customfont)
        self.session_label.grid(row=4, column=0, padx=20, pady=10, sticky='e')

        optionmenu_var = customtkinter.StringVar(value="Forenoon")
        self.session_entry = customtkinter.CTkOptionMenu(master=self.form_frame, values=[
                                                         "Forenoon", "Afternoon", "AllDay", "Other"], variable=optionmenu_var, width=200)
        self.session_entry.grid(row=4, column=1, padx=20, pady=10)

        self.hours_label = customtkinter.CTkLabel(
            master=self.form_frame, text='Hours: ', compound='right', anchor='w', font=self.customfont)
        self.hours_label.grid(row=5, column=0, padx=20, pady=10, sticky='e')

        self.hours_entry = customtkinter.CTkEntry(
            master=self.form_frame, placeholder_text='2-5', width=200)
        self.hours_entry.grid(row=5, column=1, padx=20, pady=10)

        self.remarks_label = customtkinter.CTkLabel(
            master=self.form_frame, text='Remarks: ', compound='right', anchor='w', font=self.customfont)
        self.remarks_label.grid(row=6, column=0, padx=20, pady=10, sticky='e')

        self.remarks_entry = customtkinter.CTkEntry(
            master=self.form_frame, placeholder_text='remarks', width=200)
        self.remarks_entry.grid(row=6, column=1, padx=20, pady=10)

        self.list_items = [f"{item[0]}   {item[1]}" for item in self.excel_data]
        self.listbox = ScrollableCheckBoxFrame(self.list_frame, label_text="List of students", width=350, height=500, item_list=self.list_items, label_fg_color='#00ADFE')
        self.listbox.grid(row=0, column=0, padx=10, pady=30,columnspan=2)

        self.sel_listbox = ScrollableCheckBoxFrame1(
            self.list_frame, width=350, height=500, label_text="Selected students", label_fg_color='#00ADFE')
        self.sel_listbox.grid(row=0, column=2, padx=10, pady=10,columnspan=2)

        self.students_label = customtkinter.CTkLabel(
            self.list_frame, text=f'Total count:   {len(self.list_items)}')
        self.students_label.grid(row=1, column=0,columnspan=2)

        self.selected_label = customtkinter.CTkLabel(
            self.list_frame, text=f'Total count:   {len(self.sel_listbox.get_checked_items())}')
        self.selected_label.grid(row=1, column=2, columnspan=2)

        # self.select_all = customtkinter.CTkButton(
        #     self.list_frame, text='Select all', height=32, command=self.test)
        # self.select_all.grid(row=2, column=0, padx=10, pady=10)

        self.push_selected = customtkinter.CTkButton(
            self.list_frame, text='Confirm', height=32, command=self.push_selected)
        self.push_selected.grid(
            row=2, column=0, padx=10, pady=10, columnspan=2)

        self.clear_all = customtkinter.CTkButton(
            self.list_frame, text='Clear all', height=32, command=self.reset_all)
        self.clear_all.grid(row=2, column=2, padx=10, pady=10)

        self.save = customtkinter.CTkButton(
            self.list_frame, text='Save', fg_color='green', height=32, command=self.save_all)
        self.save.grid(row=2, column=3, padx=10, pady=10)

    def reset_all(self):
        students = self.sel_listbox.get_checked_items()
        if students:
            for student in students:
                self.sel_listbox.remove_item(f'{student}')
                self.selected_label.configure(
                    text=f'Total count:   {len(self.sel_listbox.get_checked_items())}')

    def push_selected(self):
        students = self.listbox.get_checked_items()
        if students:
            for student in students:
                self.sel_listbox.add_item(f'{student}')
                self.selected_label.configure(text=f'Total count:   {len(self.sel_listbox.get_checked_items())}')

    # def test(self):
    #     self.listbox.add_item(item=self.list_items,all=True)

    def save_all(self):
        students = self.sel_listbox.get_checked_items()
        reg = []
        stu = []
        for i in range(len(students)):
            splice = students[i].split('   ')[0]
            reg.append(splice)
            splice2 = students[i].split('   ')[1]
            stu.append(splice2)

        result = list(zip(reg, stu))

        self.event_name = self.event_entry.get()
        self.from_date = self.from_entry.get_date()
        self.to_date = self.to_entry.get_date()
        self.session = self.session_entry.get()
        self.hours = self.hours_entry.get()
        self.remarks = self.remarks_entry.get()
        if self.session == 'Allday':
            self.hours = 'NA'

        if(self.event_name and self.session and self.hours and result):
            # Create a new workbook if it doesn't exist
            user_filename = filedialog.asksaveasfilename(
                defaultextension='.xlsx')
            if user_filename:
                try:
                    wbf = load_workbook(user_filename)
                except:
                    wbf = Workbook()

                wsf = wbf.active
                wsf['A1'] = 'Event Name'
                wsf['B1'] = 'From Date'
                wsf['C1'] = 'To Date'
                wsf['D1'] = 'Reg.No'
                wsf['E1'] = 'Student Name'
                wsf['F1'] = 'Session'
                wsf['G1'] = 'No. of Hours'
                wsf['H1'] = 'Remarks'
                alignment = Alignment(wrap_text=True,horizontal='center',vertical='center')

                # Append data to the next empty row
                for i, (key, value) in enumerate(result):
                    wsf.append([self.event_name, self.from_date, self.to_date,
                                key, value, self.session, self.hours, self.remarks])
                    wbf.save(user_filename)
                showinfo(
                    title='Succuss',
                    message='Data has been saved successfully!')
            else:
                showerror(title='Error', message="File couldn't be saved")
        else:
            showerror(title='Error', message='Please fill all the Entries')

# home = HomePage()
# home.mainloop()
